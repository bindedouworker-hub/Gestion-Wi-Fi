"""Report service — generates PDF and Excel reports."""

from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from sqlalchemy.orm import Session
from sqlalchemy import func, and_

from app.models.sale import Sale
from app.models.ticket import Ticket, TicketStatus
from app.models.user import User


def get_sales_data(
    db: Session,
    start_date: date | None = None,
    end_date: date | None = None,
    vendor_id: int | None = None,
) -> list[dict]:
    """Get sales data for report generation."""
    query = (
        db.query(Sale, Ticket, User)
        .join(Ticket, Sale.ticket_id == Ticket.id)
        .join(User, Sale.vendor_id == User.id)
    )

    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
        query = query.filter(Sale.created_at >= start_dt)
    if end_date:
        end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)
        query = query.filter(Sale.created_at <= end_dt)
    if vendor_id:
        query = query.filter(Sale.vendor_id == vendor_id)

    results = query.order_by(Sale.created_at.desc()).all()

    return [
        {
            "id": sale.id,
            "date": sale.created_at.strftime("%d/%m/%Y %H:%M"),
            "ticket_code": ticket.code,
            "vendor": user.full_name,
            "client_name": sale.client_name or "-",
            "client_phone": sale.client_phone or "-",
            "payment_method": sale.payment_method,
            "amount": float(sale.amount),
            "cancelled": "Oui" if sale.is_cancelled else "Non",
        }
        for sale, ticket, user in results
    ]


def get_report_dates(report_type: str) -> tuple[date, date]:
    """Calculate start and end dates based on report type."""
    today = date.today()
    if report_type == "daily":
        return today, today
    elif report_type == "weekly":
        start = today - timedelta(days=today.weekday())
        return start, today
    elif report_type == "monthly":
        start = today.replace(day=1)
        return start, today
    else:
        return today, today


def generate_pdf_report(sales_data: list[dict], title: str) -> BytesIO:
    """Generate a PDF report from sales data."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Spacer(1, 20))

    # Summary
    total_sales = len([s for s in sales_data if s["cancelled"] == "Non"])
    total_revenue = sum(s["amount"] for s in sales_data if s["cancelled"] == "Non")
    cash_total = sum(s["amount"] for s in sales_data if s["payment_method"] == "cash" and s["cancelled"] == "Non")
    wave_total = sum(s["amount"] for s in sales_data if s["payment_method"] != "cash" and s["cancelled"] == "Non")

    summary = f"Ventes: {total_sales} | Chiffre d'affaires: {total_revenue:,.0f} FCFA | Espèces: {cash_total:,.0f} FCFA | Wave: {wave_total:,.0f} FCFA"
    elements.append(Paragraph(summary, styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Table
    if sales_data:
        headers = ["#", "Date", "Code", "Vendeur", "Client", "Téléphone", "Paiement", "Montant", "Annulé"]
        table_data = [headers]
        for s in sales_data:
            table_data.append([
                s["id"], s["date"], s["ticket_code"], s["vendor"],
                s["client_name"], s["client_phone"], s["payment_method"],
                f"{s['amount']:,.0f}", s["cancelled"],
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Aucune vente pour cette période.", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(sales_data: list[dict], title: str) -> BytesIO:
    """Generate an Excel report from sales data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport des ventes"

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"].value = title
    ws["A1"].font = Font(size=14, bold=True, color="1e3a5f")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary row
    total_sales = len([s for s in sales_data if s["cancelled"] == "Non"])
    total_revenue = sum(s["amount"] for s in sales_data if s["cancelled"] == "Non")
    ws.merge_cells("A2:I2")
    ws["A2"].value = f"Ventes: {total_sales} | Chiffre d'affaires: {total_revenue:,.0f} FCFA"
    ws["A2"].font = Font(size=10, italic=True)

    # Headers
    headers = ["#", "Date", "Code", "Vendeur", "Client", "Téléphone", "Paiement", "Montant (FCFA)", "Annulé"]
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for row_idx, s in enumerate(sales_data, 5):
        values = [s["id"], s["date"], s["ticket_code"], s["vendor"],
                  s["client_name"], s["client_phone"], s["payment_method"],
                  s["amount"], s["cancelled"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
